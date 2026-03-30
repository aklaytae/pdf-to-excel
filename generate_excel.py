#!/usr/bin/env python3
import sys
import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

def thin_border():
    s = Side(style='thin', color='D0D7E0')
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thick = Side(style='medium', color='1F4E79')
    thin  = Side(style='thin',   color='D0D7E0')
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def generate_excel(data, output_path):
    meta         = data['meta']
    transactions = data['transactions']

    # ── Aggregate by date ────────────────────────────────────────────────────
    daily = defaultdict(lambda: {'debit': 0.0, 'credit': 0.0, 'count': 0})
    for t in transactions:
        daily[t['date']]['debit']  += t['debit']
        daily[t['date']]['credit'] += t['credit']
        daily[t['date']]['count']  += 1

    # ── Find full date range from first to last date ─────────────────────────
    def to_ce(d):
        p = d.split('/')
        return datetime.date(int(p[2]) + 2500 - 543, int(p[1]), int(p[0]))

    all_dates = sorted(daily.keys(), key=to_ce)
    start_dt  = to_ce(all_dates[0])
    end_dt    = to_ce(all_dates[-1])

    # Build full date list (every calendar day, no gaps)
    full_dates = []
    cur = start_dt
    while cur <= end_dt:
        full_dates.append(cur)
        cur += datetime.timedelta(days=1)

    # Day name (Thai)
    DAY_TH = {0:'จันทร์',1:'อังคาร',2:'พุธ',3:'พฤหัสบดี',4:'ศุกร์',5:'เสาร์',6:'อาทิตย์'}

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'รายรับ-รายจ่ายรายวัน'

    NAVY    = '1F4E79'
    TEAL    = '2E86AB'
    RED_BG  = 'FDECEA'
    GRN_BG  = 'E8F5E9'
    ZERO_BG = 'F5F5F5'
    RED_FG  = 'C0392B'
    GRN_FG  = '1E8449'
    MUT_FG  = '9E9E9E'
    WHT     = 'FFFFFF'

    hdr_font  = Font(name='TH Sarabun New', bold=True, size=12, color=WHT)
    hdr_fill  = PatternFill('solid', start_color=NAVY)
    ctr       = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = f"สรุปรายรับ-รายจ่ายรายวัน"
    c.font  = Font(name='TH Sarabun New', bold=True, size=15, color=WHT)
    c.fill  = PatternFill('solid', start_color=NAVY)
    c.alignment = ctr
    ws.row_dimensions[1].height = 32

    # ── Row 2: Account info ───────────────────────────────────────────────────
    ws.merge_cells('A2:G2')
    c = ws['A2']
    name = meta.get('account_name', '')
    acc  = meta.get('account_number', '')
    start_str = f"{start_dt.day:02d}/{start_dt.month:02d}/{start_dt.year+543}"
    end_str   = f"{end_dt.day:02d}/{end_dt.month:02d}/{end_dt.year+543}"
    c.value = f"ชื่อบัญชี: {name}  |  เลขที่บัญชี: {acc}  |  ช่วง: {start_str} – {end_str}"
    c.font  = Font(name='TH Sarabun New', size=11, color=WHT)
    c.fill  = PatternFill('solid', start_color=TEAL)
    c.alignment = ctr
    ws.row_dimensions[2].height = 22

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    COLS = ['วันที่', 'วัน', 'จำนวนรายการ', 'รายจ่าย (บาท)', 'รายรับ (บาท)', 'ยอดสุทธิ (บาท)', 'หมายเหตุ']
    for col, h in enumerate(COLS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = ctr
        c.border    = thick_bottom()
    ws.row_dimensions[3].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    DATA_START = 4
    for r_off, dt in enumerate(full_dates):
        row = DATA_START + r_off

        # Lookup key in format DD/MM/YY (Thai year 2-digit)
        yr_th2 = str(dt.year + 543 - 2500)   # 2025 -> 68
        key = f"{dt.day:02d}/{dt.month:02d}/{yr_th2}"
        d   = daily.get(key)

        debit  = d['debit']   if d else 0.0
        credit = d['credit']  if d else 0.0
        count  = d['count']   if d else 0
        net    = credit - debit
        no_tx  = (count == 0)

        # Date string in Thai Buddhist year
        yr_be  = dt.year + 543        # 2025 -> 2568
        date_s = f"{dt.day:02d}/{dt.month:02d}/{yr_be}"
        day_s  = DAY_TH[dt.weekday()]

        # Row fill
        if no_tx:
            fill = PatternFill('solid', start_color=ZERO_BG)
        elif net < 0:
            fill = PatternFill('solid', start_color=RED_BG)
        else:
            fill = PatternFill('solid', start_color=GRN_BG)

        note = 'ไม่มีรายการ' if no_tx else ''

        row_data = [date_s, day_s, count, debit, credit, net, note]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill   = fill
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal='center' if col in (1,2,3,7) else 'right',
                vertical='center'
            )

            if no_tx:
                c.font = Font(name='TH Sarabun New', size=11, color=MUT_FG)
            elif col == 4:   # debit
                c.font = Font(name='TH Sarabun New', size=11, color=RED_FG)
            elif col == 5:   # credit
                c.font = Font(name='TH Sarabun New', size=11, color=GRN_FG)
            elif col == 6:   # net
                color = RED_FG if net < 0 else GRN_FG
                c.font = Font(name='TH Sarabun New', bold=True, size=11, color=color)
            else:
                c.font = Font(name='TH Sarabun New', size=11)

            if col in (4, 5, 6):
                c.number_format = '#,##0.00'

        ws.row_dimensions[row].height = 20

    # ── Total row ─────────────────────────────────────────────────────────────
    n     = len(full_dates)
    t_row = DATA_START + n
    ws.merge_cells(f'A{t_row}:C{t_row}')

    total_fill = PatternFill('solid', start_color='2C3E50')
    total_font = Font(name='TH Sarabun New', bold=True, size=12, color=WHT)

    c = ws['A' + str(t_row)]
    c.value     = f'รวมทั้งหมด  ({len(full_dates)} วัน)'
    c.font      = total_font
    c.fill      = total_fill
    c.alignment = ctr
    c.border    = thin_border()

    for col in range(1, 8):
        tc = ws.cell(row=t_row, column=col)
        tc.fill   = total_fill
        tc.border = thin_border()
        tc.alignment = Alignment(horizontal='center' if col in (1,2,3) else 'right', vertical='center')

    d_total = ws.cell(row=t_row, column=4,
                      value=f'=SUM(D{DATA_START}:D{t_row-1})')
    d_total.font = Font(name='TH Sarabun New', bold=True, size=12, color='FF6B6B')
    d_total.number_format = '#,##0.00'
    d_total.fill = total_fill; d_total.border = thin_border()
    d_total.alignment = Alignment(horizontal='right', vertical='center')

    c_total = ws.cell(row=t_row, column=5,
                      value=f'=SUM(E{DATA_START}:E{t_row-1})')
    c_total.font = Font(name='TH Sarabun New', bold=True, size=12, color='55EFC4')
    c_total.number_format = '#,##0.00'
    c_total.fill = total_fill; c_total.border = thin_border()
    c_total.alignment = Alignment(horizontal='right', vertical='center')

    n_total = ws.cell(row=t_row, column=6,
                      value=f'=E{t_row}-D{t_row}')
    n_total.font = Font(name='TH Sarabun New', bold=True, size=12, color='FFEAA7')
    n_total.number_format = '#,##0.00'
    n_total.fill = total_fill; n_total.border = thin_border()
    n_total.alignment = Alignment(horizontal='right', vertical='center')

    ws.row_dimensions[t_row].height = 28

    # ── Freeze panes & column widths ─────────────────────────────────────────
    ws.freeze_panes = 'A4'  # freeze header rows

    ws.column_dimensions['A'].width = 15   # วันที่
    ws.column_dimensions['B'].width = 14   # วัน
    ws.column_dimensions['C'].width = 16   # จำนวนรายการ
    ws.column_dimensions['D'].width = 18   # รายจ่าย
    ws.column_dimensions['E'].width = 18   # รายรับ
    ws.column_dimensions['F'].width = 20   # ยอดสุทธิ
    ws.column_dimensions['G'].width = 14   # หมายเหตุ

    # Auto-filter
    ws.auto_filter.ref = f'A3:G{t_row}'

    wb.save(output_path)
    return {
        'sheets': 1,
        'months': [],
        'total_days': len(full_dates),
        'tx_days': sum(1 for dt in full_dates
                       if f"{dt.day:02d}/{dt.month:02d}/{str(dt.year+543-2500)}" in daily),
        'zero_days': sum(1 for dt in full_dates
                         if f"{dt.day:02d}/{dt.month:02d}/{str(dt.year+543-2500)}" not in daily),
    }

if __name__ == '__main__':
    with open(sys.argv[1], encoding='utf-8') as f:
        data = json.load(f)
    result = generate_excel(data, sys.argv[2])
    print(json.dumps(result, ensure_ascii=False))
